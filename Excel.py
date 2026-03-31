import pandas as pd
import openpyxl
import re
from copy import copy


def extract_iin(text):
    if pd.isna(text): return None
    match = re.search(r'\b\d{12}\b', str(text))
    return match.group(0) if match else None


def main():
    print("1. Загрузка базы ИИН из Excel2...")
    valid_iins = set()
    xls2 = pd.ExcelFile('Excel2.xlsx')
    for sheet in xls2.sheet_names:
        df2 = pd.read_excel(xls2, sheet_name=sheet)
        if df2.shape[1] > 1:
            valid_iins.update(df2.iloc[:, 1].apply(extract_iin).dropna().tolist())

    print("2. Открытие оригинала Excel1...")
    wb = openpyxl.load_workbook('Excel1.xlsx')
    ws = wb.active

    # Запоминаем данные и стили строк (начиная с 8-й строки)
    print("3. Анализ строк и стилей...")
    rows_to_keep = []
    # Проходим по строкам данных (с 8 по самую последнюю)
    for row in ws.iter_rows(min_row=8):
        # ИИН находится в 9-й колонке (индекс 8)
        iin_value = row[8].value
        iin = extract_iin(iin_value)

        if iin in valid_iins:
            rows_to_keep.append(row)

    print(f"   Найдено совпадений: {len(rows_to_keep)}. Очистка и перезапись...")

    # Удаляем все старые строки данных одним махом (это быстро)
    ws.delete_rows(8, ws.max_row)

    # Записываем только те, что нужно сохранить
    for row_idx, old_row in enumerate(rows_to_keep, start=8):
        # Устанавливаем высоту строки как в оригинале
        ws.row_dimensions[row_idx].height = ws.row_dimensions[old_row[0].row].height

        for col_idx, cell in enumerate(old_row, start=1):
            new_cell = ws.cell(row=row_idx, column=col_idx, value=cell.value)

            # Копируем стиль ячейки (шрифт, цвет, границы, выравнивание)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    output = "Excel1_Финал_Стиль.xlsx"
    print(f"4. Сохранение в {output}...")
    wb.save(output)
    print("ГОТОВО! Проверь результат.")


if __name__ == "__main__":
    main()